import hashlib
from io import BytesIO
from typing import Tuple

from openpyxl import load_workbook


def get_size_and_hash(file_to_upload: str, file_format: str) -> Tuple[int, str]:
    """Return the size and hash of file to upload

    Args:
        file_to_upload: File to upload
        file_format (str): File format

    Returns:
        Tuple[int, str]: Tuple (size, hash)
    """
    f = open(file_to_upload, "rb")
    md5hash = hashlib.md5()
    if file_format == "xlsx":
        first_chunk = f.read(4096)
        size = len(first_chunk)
        signature = first_chunk[:4]
        if signature == b"PK\x03\x04":  # xlsx
            xlsxbuffer = bytearray(first_chunk)
            while chunk := f.read(4096):
                size += len(chunk)
                xlsxbuffer.extend(chunk)
            workbook = load_workbook(filename=BytesIO(xlsxbuffer), read_only=True)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for cols in sheet.iter_rows(values_only=True):
                    md5hash.update(bytes(str(cols), "utf-8"))
            workbook.close()
            return size, md5hash.hexdigest()
        else:
            md5hash.update(first_chunk)
    else:
        size = 0
    while chunk := f.read(4096):
        size += len(chunk)
        md5hash.update(chunk)
    return size, md5hash.hexdigest()
